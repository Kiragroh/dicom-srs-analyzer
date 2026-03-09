"""
SRS Analysis Pipeline - Excel I/O
===================================
Builds, updates and reads the PTV mapping Excel table.
Existing manual entries (ExcludeFromAnalysis, Prescription_Gy_reference,
NewStructureName, Comment) are never overwritten.
"""

import logging
import os
from typing import List, Optional, Dict

import numpy as np
import pandas as pd

from config import MAPPING_EXCEL_PATH, ANONYMIZE_OUTPUT
from structure_mapping import _is_excluded
from anonymization import anonymize_patient_id, anonymize_structure_name

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

MAPPING_COLUMNS = [
    "PatientFolder",
    "PatientID",
    "PlanType",                # info only; mapping shared across HA/MM
    "PlanLabel",
    "StructureName_Original",
    "Volume_cc",
    "IsPTV_candidate_auto",
    "ExcludeFromAnalysis",     # EDITABLE: True/False
    "ExcludeReason",           # EDITABLE: reason text
    "Prescription_Gy_detected", # READ-ONLY: auto-detected from DVH
    "Prescription_Gy_reference", # EDITABLE: override Rx here → auto-renames PTV##_##Gy
    "NewStructureName",        # EDITABLE: custom name (or leave for auto PTV##_##Gy)
    "Comment",                 # EDITABLE: notes
]

# Columns that may only be set manually – never overwritten by the pipeline
MANUAL_COLUMNS = {
    "ExcludeFromAnalysis",
    "ExcludeReason",
    "NewStructureName",
    "Comment",
}
# Note: Prescription_Gy_reference is AUTO-POPULATED from detected value on first
# run ("update if empty" rule), so users can edit it directly without knowing the
# detected value.  NewStructureName auto-names (PTV##_##Gy) are regenerated
# automatically when this value changes.

# Unique key identifying a row (shared across HA/MM plan types)
KEY_COLS = ["PatientFolder", "StructureName_Original"]


# ---------------------------------------------------------------------------
# Load existing mapping
# ---------------------------------------------------------------------------

def load_mapping(path: str = MAPPING_EXCEL_PATH) -> pd.DataFrame:
    """
    Load the existing mapping Excel.  Returns an empty DataFrame with the
    correct columns if the file does not exist yet.
    """
    if not os.path.isfile(path):
        logger.info("No existing mapping file at %s – starting fresh.", path)
        return pd.DataFrame(columns=MAPPING_COLUMNS)

    try:
        df = pd.read_excel(path, dtype=str, keep_default_na=False)
        logger.info("Loaded existing mapping: %d rows from %s", len(df), path)
        # Ensure all expected columns exist
        for col in MAPPING_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[MAPPING_COLUMNS]
    except Exception as exc:
        logger.error("Failed to read mapping file %s: %s", path, exc)
        return pd.DataFrame(columns=MAPPING_COLUMNS)


# ---------------------------------------------------------------------------
# Save mapping
# ---------------------------------------------------------------------------

def save_mapping(df: pd.DataFrame, path: str = MAPPING_EXCEL_PATH) -> None:
    """Save the mapping DataFrame to Excel with column highlighting."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        if ANONYMIZE_OUTPUT:
            df_out = df.copy()
            if 'PatientFolder' in df_out.columns:
                df_out['PatientFolder'] = df_out['PatientFolder'].apply(anonymize_patient_id)
            if 'PatientID' in df_out.columns:
                df_out['PatientID'] = df_out['PatientID'].apply(anonymize_patient_id)
            if 'StructureName_Original' in df_out.columns:
                df_out['StructureName_Original'] = df_out['StructureName_Original'].apply(anonymize_structure_name)
            if 'NewStructureName' in df_out.columns:
                df_out['NewStructureName'] = df_out['NewStructureName'].apply(lambda x: anonymize_structure_name(x) if x else x)
        else:
            df_out = df

        df_out.to_excel(path, index=False)

        # Highlight editable columns with openpyxl
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            wb = load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            highlight_cols = {
                "Prescription_Gy_reference": "FFF2CC",  # yellow – edit to override Rx
                "ExcludeFromAnalysis":        "FFE0E0",  # light red
                "NewStructureName":            "E0F0FF",  # light blue
                "Comment":                     "F0F0F0",  # light gray
            }
            for col_name, color in highlight_cols.items():
                if col_name in headers:
                    ci = headers.index(col_name) + 1
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    ws.cell(row=1, column=ci).font = Font(bold=True)
                    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                        for cell in row:
                            cell.fill = fill
            wb.save(path)
        except Exception as fmt_exc:
            logger.debug("Excel formatting skipped: %s", fmt_exc)

        logger.info("Saved mapping: %d rows → %s", len(df), path)
    except Exception as exc:
        logger.error("Failed to save mapping to %s: %s", path, exc)


# ---------------------------------------------------------------------------
# Merge / update helper
# ---------------------------------------------------------------------------

def upsert_mapping(
    existing_df: pd.DataFrame,
    new_rows: List[dict],
) -> pd.DataFrame:
    """
    Merge *new_rows* into *existing_df*.

    Rules:
    - If a row (same KEY_COLS) already exists → preserve manual columns;
      update auto-detected columns only if they are currently empty.
    - If a row is new → append it.
    - Log each action (new / updated / unchanged).
    """
    if existing_df.empty:
        result = pd.DataFrame(new_rows, columns=MAPPING_COLUMNS).fillna("")
        logger.info("  Created %d new rows (empty mapping).", len(result))
        return result

    # Build index on existing rows
    def make_key(row):
        return (
            str(row.get("PatientFolder", "") or ""),
            str(row.get("StructureName_Original", "") or ""),
        )

    existing_index: Dict[tuple, int] = {
        make_key(existing_df.iloc[i]): i
        for i in range(len(existing_df))
    }

    updated_df = existing_df.copy()
    new_rows_to_append = []
    counts = {"new": 0, "updated": 0, "unchanged": 0, "manual_preserved": 0}

    for row in new_rows:
        key = make_key(row)
        if key in existing_index:
            idx = existing_index[key]
            changed = False
            for col in MAPPING_COLUMNS:
                if col in MANUAL_COLUMNS:
                    # Never overwrite manual columns
                    existing_val = str(updated_df.at[idx, col]).strip()
                    if existing_val:
                        counts["manual_preserved"] += 1
                    continue
                # SPECIAL: Prescription_Gy_reference is editable → preserve if set
                if col == "Prescription_Gy_reference":
                    existing_val = str(updated_df.at[idx, col]).strip()
                    if existing_val:
                        counts["manual_preserved"] += 1
                        continue
                # Overwrite auto columns only if currently empty
                existing_val = str(updated_df.at[idx, col]).strip()
                new_val = str(row.get(col, "") or "").strip()
                if not existing_val and new_val:
                    updated_df.at[idx, col] = new_val
                    changed = True
            if changed:
                counts["updated"] += 1
            else:
                counts["unchanged"] += 1
        else:
            # Brand-new row – append
            new_rows_to_append.append(row)
            counts["new"] += 1

    if new_rows_to_append:
        append_df = pd.DataFrame(new_rows_to_append, columns=MAPPING_COLUMNS).fillna("")
        updated_df = pd.concat([updated_df, append_df], ignore_index=True)

    logger.info(
        "  Mapping merge: %d new | %d updated | %d unchanged | %d manual columns preserved",
        counts["new"], counts["updated"], counts["unchanged"], counts["manual_preserved"],
    )
    return updated_df


# ---------------------------------------------------------------------------
# Build rows from loaded plan data
# ---------------------------------------------------------------------------

def build_mapping_rows(
    plan_files,          # PlanFiles
    structures: list,    # List[Structure]
    dose,                # Optional[DoseData]
    plan_meta,           # Optional[PlanMeta]
) -> List[dict]:
    """
    Create one mapping row per structure detected as a PTV candidate.
    Auto-columns are filled; manual columns are left empty (they will be
    preserved if they already exist in the file).
    """
    from structure_mapping import is_ptv_candidate, estimate_prescription_from_dvh

    rows = []
    patient_id = plan_meta.patient_id if plan_meta else "Unknown"
    plan_label = plan_meta.plan_label if plan_meta else "Unknown"

    for struct in structures:
        is_candidate = is_ptv_candidate(struct)
        if not is_candidate:
            continue   # Only include PTV candidates in the mapping table

        # Estimate volume (cc) from contour areas  – bounding-box voxel method
        vol_cc = _estimate_volume_cc(struct, dose)

        # Estimate prescription dose
        rx_detected = None
        if dose is not None:
            rx_detected = estimate_prescription_from_dvh(struct, dose)

        row = {col: "" for col in MAPPING_COLUMNS}
        row.update({
            "PatientFolder":           plan_files.patient_folder,
            "PatientID":               patient_id,
            "PlanType":                plan_files.plan_type,
            "PlanLabel":               plan_label,
            "StructureName_Original":  struct.name,
            "Volume_cc":               f"{vol_cc:.3f}" if vol_cc is not None else "",
            "IsPTV_candidate_auto":    "True",
            "ExcludeFromAnalysis":     "",   # manual column – left empty
            "ExcludeReason":           "",   # manual
            "Prescription_Gy_detected":  str(int(rx_detected)) if rx_detected else "",
            "Prescription_Gy_reference":  str(int(rx_detected)) if rx_detected else "",  # pre-populated; user may override
            "NewStructureName":        "",   # assigned later after merge
            "Comment":                 "",   # manual
        })
        rows.append(row)
        logger.debug(
            "  Candidate: %s/%s/%s  Rx_det=%s",
            plan_files.patient_folder, plan_files.plan_type,
            struct.name, rx_detected,
        )

    logger.info(
        "  Built %d PTV candidate rows for %s/%s",
        len(rows), plan_files.patient_folder, plan_files.plan_type,
    )
    return rows


def _estimate_volume_cc(structure, dose) -> Optional[float]:
    """
    Rough volume estimate for the structure (in cc).
    Uses a 2 mm grid if dose is available, otherwise returns None.
    """
    from structure_mapping import (
        _get_bounding_box_points, _point_in_structure_xy,
    )
    if not structure.contours:
        return None
    step = 2.0
    pts = _get_bounding_box_points(structure, step)
    if not pts:
        return None
    inside = sum(1 for p in pts if _point_in_structure_xy(p, structure))
    vol_mm3 = inside * (step ** 3)
    return vol_mm3 / 1000.0  # cm³ = cc


# ---------------------------------------------------------------------------
# Post-merge: assign NewStructureName per plan
# ---------------------------------------------------------------------------

def _is_auto_name(name: str) -> bool:
    """Return True if name looks like an auto-generated PTV##_##Gy label."""
    import re
    return bool(re.fullmatch(r'PTV\d+_\d+Gy', str(name).strip()))


def _rx_in_auto_name(name: str):
    """Extract dose from auto-name PTV##_##Gy; return int or None."""
    import re
    m = re.fullmatch(r'PTV\d+_(\d+)Gy', str(name).strip())
    return int(m.group(1)) if m else None


def assign_new_names_inplace(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each PatientFolder group (shared across HA/MM), assign NewStructureName
    to rows where it is currently empty and not excluded.

    Auto-generated names (pattern PTV##_##Gy) are cleared and regenerated
    when the embedded dose no longer matches Prescription_Gy_reference,
    so users only need to edit that column – no need to also update the name.
    Manually customised names (not matching the auto-pattern) are always kept.
    """
    from structure_mapping import build_new_structure_names

    groups = df.groupby(["PatientFolder"])
    for folder, group_idx in groups.groups.items():
        # Step 1 – clear all auto-generated names so Step 2 can renumber
        # them from scratch (handles Rx changes, exclusion changes, etc.).
        # Manually customised names (not matching PTV##_##Gy) are never touched.
        for i in group_idx:
            cur_name = str(df.at[i, "NewStructureName"]).strip()
            if _is_auto_name(cur_name):
                df.at[i, "NewStructureName"] = ""

        # Step 2 – regenerate ALL auto-names for this patient (to ensure consistent numbering)
        rows_list = df.loc[group_idx].to_dict("records")
        
        # Collect all active PTV rows (including those with stale names cleared above)
        active_rows = [
            r for r in rows_list
            if str(r.get("IsPTV_candidate_auto", "")).lower() == "true"
            and not _is_excluded(r)
        ]
        
        if not active_rows:
            continue
        
        # Regenerate names for all active PTVs (ensures counter starts at 1 per patient)
        build_new_structure_names(active_rows)

        # Apply regenerated names back to dataframe
        name_map = {
            r["StructureName_Original"]: r["NewStructureName"]
            for r in active_rows
        }
        for i in group_idx:
            orig = df.at[i, "StructureName_Original"]
            if orig in name_map and not str(df.at[i, "NewStructureName"]).strip():
                df.at[i, "NewStructureName"] = name_map[orig]

    return df


# ---------------------------------------------------------------------------
# Exclusion report
# ---------------------------------------------------------------------------

def get_excluded_structures(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame of rows that are marked as excluded."""
    mask = df["ExcludeFromAnalysis"].apply(
        lambda v: str(v).strip().lower() in {"true", "1", "ja", "yes", "x"}
    )
    excluded = df[mask].copy()
    if not excluded.empty:
        logger.info("Excluded structures (%d total):", len(excluded))
        for _, row in excluded.iterrows():
            logger.info(
                "  EXCLUDED: %s/%s/%s  reason=%s",
                row["PatientFolder"], row["PlanType"],
                row["StructureName_Original"],
                row.get("ExcludeReason", ""),
            )
    return excluded


def get_active_structures(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame of rows that are NOT excluded and ARE PTV candidates."""
    is_candidate = df["IsPTV_candidate_auto"].astype(str).str.lower() == "true"
    is_excluded_mask = df["ExcludeFromAnalysis"].apply(
        lambda v: str(v).strip().lower() in {"true", "1", "ja", "yes", "x"}
    )
    return df[is_candidate & ~is_excluded_mask].copy()
